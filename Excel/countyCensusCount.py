import openpyxl
import pprint

print('Opening workbook...')
census = openpyxl.load_workbook('censuspopdata.xlsx')

sheet = census.active

countyData = {}

# TODO: Determine spreadsheet format/dimensions

width = sheet.max_column
height = sheet.max_row

# information = [sheet[1][i].value for i in range(width)]
# print(information[:])
# states = [(sheet[i][1].value) for i in range(2,height,1) if (sheet[i][1].value != sheet[i-1][1].value) or i == 2]
# type(states)
# 
# @dataclass
# class Census(object):
#   information[0].value : list(str)
#   information[1].value : list(str)
#   information[2].value : list(str)
#   information[3].value : list(int)
for row in range(2,height + 1):
  state = sheet['B' + str(row)].value
  county = sheet['C' + str(row)].value
  pop = sheet['D' + str(row)].value

  # Make sure the key for this state exists
  countyData.setdefault(state, {})

  # Make sure the key for the county in this state exists
  countyData[state].setdefault(county, {'tracts':0, 'pop': 0})

  #Each row represents one census tract, so increment by one
  countyData[state][county]['tracts'] += 1

  # Increase the county pops by the pop in each tract
  countyData[state][county]['pop'] += int(pop)

# Open a new text file and write the contents of countyData to it.
print('Writing results...')
with open('census2010.py', 'w') as resultFile:
  resultFile.write('allData = ' + pprint.pformat(countyData))
print('Done.')

resultFile.close()









# TODO: Obtain list of Counties

# counties = [sheet[i][2].value for i in range(2,height,1) if (sheet[i][2].value != sheet[i-1][2].value) or i == 2]
# tracts = [sum(sheet[i][3].value) if (sheet[i][2].value == sheet[i-1][2].value or i==2) else (sum(sheet[i][3].value)) for i in range(2,height,1)]

# TODO: Dict of Lists, Key = County, Value = List[Census Tract Amt, Pop of County]


# TODO: Write Dict of Lists to a text file with the .py  extension using the pprint module



# height = tally.max_row
# width = tally.max_column
# 
# sheet = [tally.cell(row=2, column=i, value=True) for i in range(1,width+1)]
# print(type(tally.cell(row=1,column=1,value=True)))
# categories = [sheet.Cell.value for ]
# for i in range(len(sheet)):
#   print(sheet[i])
#   print(sheet[i].value)
# 
# 
# 
# 
# categories = []












# dict_rows = {tally[i]:tally[i].max_row for i in range(1,amount+1,1)}
# dict_cols = {tally[i]:tally[i].max_column for i in range(1,amount+1,1)}

# rows_cols = zip(dict_rows,dict_cols)

# dict_sizes = {tally[i]:rows_cols[[i]].items() for i in range(1,amount+1,1)}