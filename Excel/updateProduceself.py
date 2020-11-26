#! python3

import openpyxl
from openpyxl import *

produce = openpyxl.load_workbook('produceSales.xlsx')
update = openpyxl.Workbook()

food = produce.get_sheet_by_name('Sheet')
foods = update.active

prices = {
  'Celery' : 1.19,
  'Garlic' : 3.07,
  'Lemon' : 1.27
}

width = food.max_column
height = food.max_row

for column in range(1, width + 1):
  foods[str(column)] = food[str(column)]
print(foods[:])


for row in range(2, height + 1):
  foods['C' + str(row)] = food['C' + str(row)].value
  if[food['A' + str(row)]].value == prices.keys('Celery'):
    foods['B' + str(row)] = food['B' + str(row)].value
  if[food['A' + str(row)]].value == prices.keys('Garlic'):
    foods['B' + str(row)] = food['B' + str(row)].value
  if[food['A' + str(row)]].value == prices.keys('Lemon'):
    foods['B' + str(row)] = food['B' + str(row)].value
  foods['D' + str(row)] = '=ROUND(B' + str(row) + '*C' + str(row) + ',2)'
  