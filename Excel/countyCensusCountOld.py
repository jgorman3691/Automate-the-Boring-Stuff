import openpyxl
import io, pprint
import sys, os, re

# TODO: Open the spreadsheet
census = []
tally = []

census = openpyxl.load_workbook('censuspopdata.xlsx')
CL = len(census.sheetnames)
print(len(census.sheetnames))
[tally.append(census.sheetnames[i+1]) for i in range(0,CL+1,1) if CL != None]

# TODO: Determine spreadsheet format/dimensions

amount = len(tally)

dict_rows = {tally[i]:tally[i].max_row for i in range(1,amount+1,1)}
dict_cols = {tally[i]:tally[i].max_column for i in range(1,amount+1,1)}

rows_cols = zip(dict_rows,dict_cols)

dict_sizes = {tally[i]:rows_cols[[i]].items() for i in range(1,amount+1,1)}


# TODO: Obtain list of Counties

# TODO: Dict of Lists, Key = County, Value = List[Census Tract Amt, Pop of County]

# TODO: Write Dict of Lists to a text file with the .py  extension using the pprint module